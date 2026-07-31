"""
Pre-processing Lambda: watches s3://<bucket>/incoming/ for the idv_control.json
control file, validates it against what's actually sitting in incoming/, moves
the batch's files into a dated batch folder, runs them through Textract (or
reads .txt directly), and caches the extracted text + confidence in RDS
Postgres under `textract_extracted_documents`.

This lets the actual KYC/ODD agent pipeline read pre-extracted text from
RDS instead of calling Textract at process time.

Trigger:
    Any .json file landing under incoming/ (top level, not inside an
    existing batch folder) is treated as the control file trigger. Its
    content is inspected before any processing happens:
        - batch_complete == false -> skip entirely, wait for the real one
        - batch_complete == true  -> validate + process

Filename convention expected under incoming/:
    <review_id>_<party_id>_<suffix>_<NN>.<file_type>

    NOTE: this is the OPPOSITE field order from the previous convention
    (<party_id>_<review_id>_<suffix>). Confirmed with source system.

    The trailing 2-digit number (_01, _02, _03, ...) is NOT a re-upload
    version - it's just a sequential index assigned by the uploading team
    across a party's documents (e.g. passport is always _01, proof of
    address is always _02). The uploading team already guarantees exactly
    one file per document type per party/review, so this Lambda does no
    deduplication - every file under incoming/ is processed as-is.

.json (other than the control file) and .csv are ignored on purpose -
those are handled by separate lambdas. Anything else that isn't
png/jpg/jpeg/tif/tiff/pdf/txt is skipped.

Validation:
    summary.total_file_count  is compared against the actual object count
        found under incoming/ (excluding the control file itself).
    summary.total_party_count is compared against the number of distinct
        party_ids parsed out of the actual filenames in incoming/.
    Any mismatch (and any "duplicate file with no version marker" case) is
    written to `pipeline_discrepancies` with source='IDV', severity='warning'.
    Mismatches are logged but do NOT block processing (warning, not critical).

Batch folder:
    The control file no longer carries a reliable batch_identifier, and this
    Lambda does NOT create the batch folder - an earlier step in the
    pipeline already did. This Lambda just scans the bucket root for
    existing BATCH_YYYYMMDD_NNN "folders" and uses the most recent one
    (highest date, then highest NNN). Every file in the batch (plus the
    control file itself, for audit trail) is then copied from incoming/ to
    <batch_identifier>/ and the original deleted, with the batch_identifier
    appended to the filename before the extension, e.g.:
        incoming/53446_43542_Proof_of_ID.jpg
          -> BATCH_20260706_001/53446_43542_Proof_of_ID_BATCH_20260706_001.jpg

Requires a Lambda layer providing PyMuPDF (fitz), openpyxl, and psycopg2
(binary) - none of these ship in the base runtime.
"""

import json
import logging
import os
import re
import urllib.parse
from datetime import datetime, timezone
from typing import Optional

import boto3
import fitz  # PyMuPDF
import openpyxl
import psycopg2
from botocore.config import Config
from botocore.exceptions import BotoCoreError, ClientError

logger = logging.getLogger()
logger.setLevel(logging.INFO)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
INCOMING_PREFIX = "incoming/"

# Bundled directly in the deployment package next to this file, e.g.:
#   lambda_function.py
#   textract_mapping.xlsx
# so no S3 fetch is needed to load it.
TEXTRACT_MAPPING_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "textract_mapping.xlsx")


ALLOWED_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".pdf", ".txt"}
IGNORED_EXTENSIONS = {".json", ".csv"}  # handled by a different lambda

DISCREPANCY_SOURCE = "IDV"
DISCREPANCY_SEVERITY = "warning"

# Simplified document types every suffix gets normalized down to.
DOCUMENT_TYPES = ("passport", "proof_of_id", "employment", "proof_of_address")

# s3_client = boto3.client("s3", config=Config(signature_version="s3v4"))
s3_client = boto3.client("s3")
textract_client = boto3.client("textract")

# Cached across warm invocations so we don't re-fetch the mapping sheet
# on every S3 event.
_variant_map_cache: Optional[dict[str, str]] = None


class SecretFetchError(Exception):
    """Raised when the RDS connection cannot be established."""


def get_connection(connect_timeout: int = 10):
    """
    Build and return a fresh psycopg2 connection using credentials from
    Secrets Manager.
    Caller is responsible for closing the connection (use try/finally or
    a context manager at the call site).
    """
    try:
        conn = psycopg2.connect(
            host ='r-dcoe-aikycdev-rds-postgresql-tf-dev-cluster.cluster-ct4o4gwekssh.eu-west-1.rds.amazonaws.com',
            port=5432,
            dbname='aikyc',
            user='aikyc',
            password='XZulm&yh(DY>9Kg+',
            sslmode='require',
        )
    except Exception as exc:  # noqa: BLE001
        raise SecretFetchError(f"Failed to connect to RDS: {exc}") from exc
    return conn


# ---------------------------------------------------------------------------
# Table reset - each new batch wipes and reloads textract_extracted_documents
# ---------------------------------------------------------------------------
DROP_AND_CREATE_SQL = """
    DROP TABLE IF EXISTS textract_extracted_documents;

    CREATE TABLE textract_extracted_documents (
        id                BIGSERIAL PRIMARY KEY,
        party_id          VARCHAR(100)  NOT NULL,
        review_id         VARCHAR(100)  NOT NULL,
        document_type     VARCHAR(50)   NOT NULL,
        original_suffix   VARCHAR(150),
        file_key          VARCHAR(1000) NOT NULL,
        file_type         VARCHAR(10)   NOT NULL,
        content           TEXT,
        avg_confidence    NUMERIC(5,2),
        status            VARCHAR(20)   NOT NULL DEFAULT 'completed',
        error_message     TEXT,
        created_at        TIMESTAMPTZ   NOT NULL DEFAULT now(),
        updated_at        TIMESTAMPTZ   NOT NULL DEFAULT now(),
        CONSTRAINT uq_textract_party_review_doctype
            UNIQUE (party_id, review_id, document_type)
    );

    CREATE INDEX idx_textract_extracted_documents_party_review
        ON textract_extracted_documents (party_id, review_id);

    CREATE INDEX idx_textract_extracted_documents_status
        ON textract_extracted_documents (status);
"""


def _reset_table() -> None:
    """Drop and recreate textract_extracted_documents for a fresh batch load."""
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(DROP_AND_CREATE_SQL)
        logger.info("textract_extracted_documents dropped and recreated for new batch")
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# pipeline_discrepancies - cumulative log table, NOT reset per batch
# ---------------------------------------------------------------------------
CREATE_DISCREPANCIES_SQL = """
    CREATE TABLE IF NOT EXISTS pipeline_discrepancies (
        id                SERIAL PRIMARY KEY,
        batch_identifier  VARCHAR(30)   NOT NULL,
        logged_at         TIMESTAMP     NOT NULL,
        source            VARCHAR(10)   NOT NULL,   -- 'EDW' or 'IDV'
        severity          VARCHAR(10)   NOT NULL,   -- 'critical' or 'warning'
        filename          VARCHAR(255)  NOT NULL,
        party_id          VARCHAR(50),              -- NULL for batch-level issues
        issue             TEXT          NOT NULL
    );
"""


def _ensure_discrepancies_table() -> None:
    """Create pipeline_discrepancies if it doesn't already exist (never dropped)."""
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(CREATE_DISCREPANCIES_SQL)
    finally:
        conn.close()


def _log_discrepancies(batch_identifier: str, discrepancies: list[dict]) -> None:
    """Insert one row per discrepancy. All rows for this lambda are source=IDV,
    severity=warning (mismatches here are informational, not batch-blocking)."""
    if not discrepancies:
        return

    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                for disc in discrepancies:
                    cur.execute(
                        """
                        INSERT INTO pipeline_discrepancies
                            (batch_identifier, logged_at, source, severity,
                             filename, party_id, issue)
                        VALUES (%s, now(), %s, %s, %s, %s, %s)
                        """,
                        (
                            batch_identifier,
                            DISCREPANCY_SOURCE,
                            DISCREPANCY_SEVERITY,
                            disc["filename"],
                            disc.get("party_id"),
                            disc["issue"],
                        ),
                    )
        logger.warning(
            "Logged %d discrepancy(ies) for batch %s", len(discrepancies), batch_identifier
        )
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Suffix -> document_type mapping (reuses textract_mapping.xlsx)
# ---------------------------------------------------------------------------
def _load_variant_map(force_reload: bool = False) -> dict[str, str]:
    """
    Load textract_mapping.xlsx bundled next to this file and flatten it into
    a {variant_suffix_lower: simplified_document_type} lookup, e.g.
    {"passport": "passport", "intl_passport": "passport", ...}
    """
    global _variant_map_cache
    if _variant_map_cache is not None and not force_reload:
        return _variant_map_cache

    col_map_keys = {
        "passport": "passport",
        "proof_of_id": "proof_of_id",
        "employment": "employment",
        "proof_of_address": "proof_of_address",
    }

    variant_map: dict[str, str] = {}

    try:
        wb = openpyxl.load_workbook(TEXTRACT_MAPPING_PATH)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        for col_idx, header in enumerate(headers):
            if not header or header.lower() not in col_map_keys:
                continue
            doc_type = col_map_keys[header.lower()]
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[col_idx]
                if val:
                    variant_map[str(val).strip().lower().lstrip("_")] = doc_type

        # Always accept the bare document-type name itself as a variant,
        # even if the mapping sheet only lists exotic aliases.
        for doc_type in DOCUMENT_TYPES:
            variant_map.setdefault(doc_type, doc_type)

        logger.info("Loaded %d suffix variants from %s", len(variant_map), TEXTRACT_MAPPING_PATH)

    except FileNotFoundError:
        logger.exception(
            "Mapping file not found at %s, falling back to defaults", TEXTRACT_MAPPING_PATH
        )
        variant_map = {doc_type: doc_type for doc_type in DOCUMENT_TYPES}

    except (KeyError, IndexError, TypeError, ValueError):
        logger.exception("Failed to parse %s, falling back to defaults", TEXTRACT_MAPPING_PATH)
        variant_map = {doc_type: doc_type for doc_type in DOCUMENT_TYPES}

    _variant_map_cache = variant_map
    return variant_map


def _resolve_document_type(raw_suffix: str) -> Optional[str]:
    """Match a raw filename suffix (version marker already stripped) against
    the loaded variant map."""
    suffix_lower = raw_suffix.lower().lstrip("_")
    variant_map = _load_variant_map()

    if suffix_lower in variant_map:
        return variant_map[suffix_lower]

    # Fall back to substring match (mirrors TextractManager._match_variant)
    for variant, doc_type in variant_map.items():
        if variant in suffix_lower:
            return doc_type

    return None


# ---------------------------------------------------------------------------
# Filename parsing:  <review_id>_<party_id>_<suffix>.<file_type>
#
# NOTE: review_id and party_id are FLIPPED relative to the old convention.
# review_id / party_id are assumed to contain no underscores; suffix may
# (and normally does, e.g. "PASSPORT_01") - the trailing number is just a
# sequential index, not a version, so it's kept as part of the suffix and
# not parsed out separately. See module docstring.
# ---------------------------------------------------------------------------
# FILENAME_RE = re.compile(r"^(?P<review_id>[^_]+)_(?P<party_id>[^_]+)_(?P<suffix>.+)$")
FILENAME_RE = re.compile(r"^(?P<party_id>[^_]+)_(?P<review_id>[^_]+)_(?P<suffix>.+)$")

# The suffix itself is <documenttype>_<NN>, e.g. "PASSPORT_01" or
# "PROOF_OF_ADDRESS_02". This strips the trailing 2-digit sequence index so
# we can look the document type up in the Textract mapping directly, rather
# than relying on substring matching against the full "documenttype_01"
# string.
SEQUENCE_INDEX_RE = re.compile(r"^(?P<doc_type_suffix>.+)_(?P<index>\d{2})$")


def _strip_sequence_index(raw_suffix: str) -> str:
    """Strip a trailing _NN sequence index off a suffix, e.g. "PASSPORT_01"
    -> "PASSPORT". If the suffix doesn't end in _NN, it's returned unchanged
    (falls through to the substring match in _resolve_document_type)."""
    match = SEQUENCE_INDEX_RE.match(raw_suffix)
    if match:
        return match.group("doc_type_suffix")
    return raw_suffix


def _parse_filename(file_name: str) -> Optional[dict]:
    name, ext = os.path.splitext(file_name)
    ext = ext.lower()

    match = FILENAME_RE.match(name)
    if not match:
        logger.warning(
            "Filename %s does not match <party_id>_<review_id>_<suffix> pattern", file_name
        )
        return None

    review_id = match.group("review_id")
    party_id = match.group("party_id")
    suffix = match.group("suffix")

    return {
        "review_id": review_id,
        "party_id": party_id,
        "suffix": suffix,
        "file_type": ext.lstrip("."),
    }


# ---------------------------------------------------------------------------
# Textract extraction - same behavior/format as TextractManager.textract_file
# ---------------------------------------------------------------------------
def _extract_lines(bucket: str, file_key: str, file_type: str) -> list[tuple[str, float]]:
    """Return a list of (text, confidence) tuples for a document in S3."""
    if file_type == "txt":
        obj = s3_client.get_object(Bucket=bucket, Key=file_key)
        raw_text = obj["Body"].read().decode("utf-8")
        return [(line, 100.0) for line in raw_text.splitlines() if line.strip()]

    if file_type in ("png", "jpg", "jpeg", "tif", "tiff"):
        obj = s3_client.get_object(Bucket=bucket, Key=file_key)
        img_bytes = obj["Body"].read()
        response = textract_client.detect_document_text(Document={"Bytes": img_bytes})
        return [
            (block["Text"], block.get("Confidence", 0.0))
            for block in response["Blocks"]
            if block["BlockType"] == "LINE"
        ]

    if file_type == "pdf":
        obj = s3_client.get_object(Bucket=bucket, Key=file_key)
        file_bytes = obj["Body"].read()
        doc = fitz.open(stream=file_bytes, filetype="pdf")

        lines: list[tuple[str, float]] = []
        for page_num, page in enumerate(doc):
            logger.info("Extracting page %d/%d from %s", page_num + 1, doc.page_count, file_key)
            pix = page.get_pixmap(dpi=200)
            img_bytes = pix.tobytes("png")
            response = textract_client.detect_document_text(Document={"Bytes": img_bytes})
            lines.extend(
                (block["Text"], block.get("Confidence", 0.0))
                for block in response["Blocks"]
                if block["BlockType"] == "LINE"
            )
        return lines

    raise ValueError(f"Unsupported file_type for Textract: {file_type}")


def _format_content(lines: list[tuple[str, float]]) -> tuple[str, Optional[float]]:
    """Build the 'Line: "..." | Confidence: XX.X%' blob plus an average confidence."""
    if not lines:
        return "", None

    formatted = "\n".join(f'Line: "{text}" | Confidence: {conf:.1f}%' for text, conf in lines)
    avg_conf = round(sum(conf for _, conf in lines) / len(lines), 2)
    return formatted, avg_conf


# ---------------------------------------------------------------------------
# Persistence
# ---------------------------------------------------------------------------
UPSERT_SQL = """
    INSERT INTO textract_extracted_documents (
        party_id, review_id, document_type, original_suffix,
        file_key, file_type, content, avg_confidence, status, error_message, updated_at
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now())
    ON CONFLICT (party_id, review_id, document_type)
    DO UPDATE SET
        original_suffix = EXCLUDED.original_suffix,
        file_key         = EXCLUDED.file_key,
        file_type        = EXCLUDED.file_type,
        content           = EXCLUDED.content,
        avg_confidence   = EXCLUDED.avg_confidence,
        status            = EXCLUDED.status,
        error_message    = EXCLUDED.error_message,
        updated_at        = now();
"""


def _save_result(
    party_id: str,
    review_id: str,
    document_type: str,
    original_suffix: str,
    file_key: str,
    file_type: str,
    content: str,
    avg_confidence: Optional[float],
    status: str,
    error_message: Optional[str] = None,
) -> None:
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(
                    UPSERT_SQL,
                    (
                        party_id,
                        review_id,
                        document_type,
                        original_suffix,
                        file_key,
                        file_type,
                        content,
                        avg_confidence,
                        status,
                        error_message,
                    ),
                )
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Control file loading
# ---------------------------------------------------------------------------
def _load_control_file(bucket: str, key: str) -> Optional[dict]:
    """Fetch and parse the control JSON. Returns None on any read/parse error."""
    try:
        obj = s3_client.get_object(Bucket=bucket, Key=key)
        raw_text = obj["Body"].read().decode("utf-8")
        return json.loads(raw_text)
    except (ClientError, BotoCoreError, json.JSONDecodeError, UnicodeDecodeError):
        logger.exception("Failed to read/parse control file %s", key)
        return None


# ---------------------------------------------------------------------------
# Batch listing - raw objects currently sitting under incoming/
# ---------------------------------------------------------------------------
def _list_incoming_files(bucket: str, prefix: str, control_key: str) -> list[dict]:
    """List every ingestible document under incoming/, skipping the control
    file itself and any ignored/unsupported/unparseable files. Each entry is
    the parsed filename dict (review_id/party_id/suffix/file_type) plus its
    S3 key under "file_key"."""
    parsed_files: list[dict] = []
    paginator = s3_client.get_paginator("list_objects_v2")

    all_seen_keys = []

    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            all_seen_keys.append(key)
            relative_key = key[len(prefix):]

            if "/" in relative_key:
                continue

            if key == control_key:
                continue

            file_name = os.path.basename(key)

            _, ext = os.path.splitext(file_name)
            ext = ext.lower()

            if ext in IGNORED_EXTENSIONS:
                logger.info("Skipping %s, handled by a different lambda", key)
                continue

            if ext not in ALLOWED_EXTENSIONS:
                logger.info("Skipping %s, unsupported extension %s", key, ext)
                continue

            parsed = _parse_filename(file_name)
            if not parsed:
                continue

            parsed["file_key"] = key
            parsed_files.append(parsed)

    logger.info("Total raw objects under %s = %d -> %s", prefix, len(all_seen_keys), all_seen_keys)
    logger.info("Total counted (post-filkter) = %d", len(parsed_files))
    return parsed_files


# ---------------------------------------------------------------------------
# Validation against the control file's declared summary
# ---------------------------------------------------------------------------
def _validate_batch(control_data: dict, actual_files: list[dict], control_filename: str) -> list[dict]:
    """Compare control_data.summary against what's actually in incoming/.
    actual_files should be the RAW (pre-dedup) list, since the control file's
    counts include every format/version variant as a separate file entry."""
    discrepancies: list[dict] = []

    summary = control_data.get("summary", {})
    declared_total_files = summary.get("total_file_count")
    declared_total_parties = summary.get("total_party_count")

    actual_total_files = len(actual_files)
    actual_party_ids = {f["party_id"] for f in actual_files}
    actual_total_parties = len(actual_party_ids)

    logger.info(
        "declared_total_files=%s declared_total_parties=%s | "
        "actual_total_files=%s actual_total_parties=%s | actual_filenames=%s",
        declared_total_files, declared_total_parties,
        actual_total_files, actual_total_parties, sorted(os.path.basename(f["file_key"])for f in actual_files)
    )

    if declared_total_files is not None and int(declared_total_files) != actual_total_files:
        discrepancies.append({
            "filename": control_filename,
            "party_id": None,
            "issue": (
                f"summary.total_file_count={declared_total_files} but "
                f"{actual_total_files} file(s) found in incoming/"
            ),
        })

    if declared_total_parties is not None and int(declared_total_parties) != actual_total_parties:
        discrepancies.append({
            "filename": control_filename,
            "party_id": None,
            "issue": (
                f"summary.total_party_count={declared_total_parties} but "
                f"{actual_total_parties} distinct party_id(s) found in incoming/"
            ),
        })

    return discrepancies


# ---------------------------------------------------------------------------
# Batch identifier lookup - a prior Lambda (the CSV-ingestion Lambda) already
# created the batch folder; we just need to find it, not create a new one.
# ---------------------------------------------------------------------------
# Batch folders are named after the drilldown CSV's timestamp
# (YYYYMMDDHHMISS, e.g. "20260604080850") by the upstream CSV-ingestion
# Lambda's BATCH_ID_PREFIX + timestamp. If that prefix is ever changed from
# "" to something else there, this regex needs to match it.
BATCH_FOLDER_RE = re.compile(r"^(?P<timestamp>\d{14})$")


class BatchFolderNotFoundError(Exception):
    """Raised when no batch folder exists yet to move files into."""


def _find_latest_batch_identifier(bucket: str) -> str:
    """Scan the bucket root for existing BATCH_YYYYMMDD_NNN folders and return
    the most recent one - highest date first, then highest NNN within that
    date. This Lambda does NOT create the folder; an earlier step in the
    pipeline already did, and we're just locating it to move files into."""
    latest: Optional[str] = None

    paginator = s3_client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=INCOMING_PREFIX, Delimiter="/"):
        for common_prefix in page.get("CommonPrefixes", []):
            folder_name = common_prefix.get("Prefix", "").rstrip("/")
            if folder_name.startswith(INCOMING_PREFIX):
                folder_name = folder_name[len(INCOMING_PREFIX):]
            match = BATCH_FOLDER_RE.match(folder_name)
            if not match:
                continue
            candidate = match.group("timestamp")
            if latest is None or candidate > latest:
                latest = candidate

    if latest is None:
        raise BatchFolderNotFoundError(
            "No batch folder found in the bucket - expected the upstream "
            "CSV-ingestion Lambda to have created one before this lambda runs"
        )  

    logger.info("Found latest batch folder: %s", latest)
    return latest


# ---------------------------------------------------------------------------
# S3 move (copy + delete) into the batch folder, suffixing filename
# ---------------------------------------------------------------------------
def _move_to_batch_folder(bucket: str, src_key: str, batch_identifier: str, prefix: str = INCOMING_PREFIX) -> str:
    """Copy src_key into <batch_identifier>/ with the batch_identifier appended
    before the extension, then delete the original. Returns the new key."""
    file_name = os.path.basename(src_key)
    name, ext = os.path.splitext(file_name)
    new_file_name = f"{name}_{batch_identifier}{ext}"
    dest_key = f"{prefix}{batch_identifier}/{new_file_name}"

    s3_client.copy_object(Bucket=bucket, CopySource={"Bucket": bucket, "Key": src_key}, Key=dest_key)
    s3_client.delete_object(Bucket=bucket, Key=src_key)

    logger.info("Moved %s -> %s", src_key, dest_key)
    return dest_key


# ---------------------------------------------------------------------------
# Single-file processing (extract + save one row)
# ---------------------------------------------------------------------------
def _process_single_file(bucket: str, file_key: str, parsed: dict) -> dict:
    party_id = parsed["party_id"]
    review_id = parsed["review_id"]
    raw_suffix = parsed["suffix"]
    file_type = parsed["file_type"]

    doc_type_suffix = _strip_sequence_index(raw_suffix)
    document_type = _resolve_document_type(doc_type_suffix)
    if not document_type:
        logger.warning(
            "Could not resolve suffix '%s' (stripped: '%s') to a known document type for %s",
            raw_suffix, doc_type_suffix, file_key,
        )
        return {"key": file_key, "status": "skipped", "reason": "unknown suffix"}

    logger.info(
        "Processing party_id=%s review_id=%s document_type=%s (%s)",
        party_id, review_id, document_type, file_key,
    )

    try:
        lines = _extract_lines(bucket, file_key, file_type)
        content, avg_confidence = _format_content(lines)

        if not content:
            logger.warning("No text extracted from %s", file_key)
            _save_result(
                party_id, review_id, document_type, raw_suffix, file_key, file_type,
                content="", avg_confidence=None, status="failed",
                error_message="No text extracted",
            )
            return {"key": file_key, "status": "failed", "reason": "no text extracted"}

        _save_result(
            party_id, review_id, document_type, raw_suffix, file_key, file_type,
            content=content, avg_confidence=avg_confidence, status="completed",
        )
        return {"key": file_key, "status": "completed", "document_type": document_type}

    except (ClientError, BotoCoreError) as exc:
        logger.exception("AWS error processing %s", file_key)
        _save_result(
            party_id, review_id, document_type, raw_suffix, file_key, file_type,
            content="", avg_confidence=None, status="failed", error_message=str(exc),
        )
        return {"key": file_key, "status": "failed", "reason": str(exc)}

    except Exception as exc:  # pylint: disable=broad-exception-caught
        logger.exception("Unexpected error processing %s", file_key)
        _save_result(
            party_id, review_id, document_type, raw_suffix, file_key, file_type,
            content="", avg_confidence=None, status="failed", error_message=str(exc),
        )
        return {"key": file_key, "status": "failed", "reason": str(exc)}


# ---------------------------------------------------------------------------
# Batch processing - triggered once by any incoming/*.json (batch_complete=true)
# ---------------------------------------------------------------------------
def _process_batch(bucket: str, prefix: str, control_data: dict, control_key: str) -> list[dict]:
    logger.info("raw control_data = %s", json.dumps(control_data))
    print("Ensuring discrepancies table")
    _ensure_discrepancies_table()

    print("Finding latest batch folder")
    batch_identifier = _find_latest_batch_identifier(bucket)

    print("Listing incoming files")
    raw_files = _list_incoming_files(bucket, prefix, control_key)

    print("Validate discrepancies")
    control_filename = os.path.basename(control_key)
    validation_discrepancies = _validate_batch(control_data, raw_files, control_filename)

    print("Log discrepancies")
    _log_discrepancies(batch_identifier, validation_discrepancies)

    print("Reset Table")
    _reset_table()

    results = []
    print("Processing files")
    for parsed in raw_files:
        moved_key = _move_to_batch_folder(bucket, parsed["file_key"], batch_identifier, prefix)
        results.append(_process_single_file(bucket, moved_key, parsed))

    # Relocate the control file itself too, for audit trail.
    try:
        _move_to_batch_folder(bucket, control_key, batch_identifier)
    except (ClientError, BotoCoreError):
        logger.exception("Failed to relocate control file %s into batch folder", control_key)

    return results


# ---------------------------------------------------------------------------
# Lambda entry point
# ---------------------------------------------------------------------------
def lambda_handler(event, context):  # noqa: ARG001  pylint: disable=unused-argument
    results = []

    for record in event.get("Records", []):
        s3_info = record.get("s3", {})
        bucket = s3_info.get("bucket", {}).get("name")
        raw_key = s3_info.get("object", {}).get("key", "")
        key = urllib.parse.unquote(raw_key.replace("+", " "))

        if not bucket:
            logger.warning("Record missing bucket name, skipping: %s", record)
            continue

        if not key.startswith(INCOMING_PREFIX):
            logger.info("Skipping %s, not under %s", key, INCOMING_PREFIX)
            continue

        # Only top-level incoming/*.json (not files already inside a batch
        # folder) are treated as control-file triggers.
        relative_key = key[len(INCOMING_PREFIX):]
        file_name = os.path.basename(key)

        if "/" in relative_key or not file_name.lower().endswith(".json"):
            logger.info(
                "Skipping %s - this lambda only triggers on a top-level .json "
                "control file landing in incoming/; individual document "
                "uploads are ignored until the control file arrives.",
                key,
            )
            continue

        print("1. Loading control file")
        control_data = _load_control_file(bucket, key)
        print("2. Control file loaded.")
        if control_data is None:
            logger.error("Could not read/parse %s - aborting for this event", key)
            continue

        if not control_data.get("batch_complete", False):
            logger.info(
                "batch_complete=false in %s - waiting for the completed control file, "
                "not processing yet",
                key,
            )
            continue

        logger.info("%s detected with batch_complete=true - validating and ingesting batch", key)
        print("3. Processing batch...")
        results.extend(_process_batch(bucket, INCOMING_PREFIX, control_data, key))
        print("4. Batch processed.")

    return {"statusCode": 200, "body": json.dumps(results)}